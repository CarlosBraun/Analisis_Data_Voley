import pdfplumber
import pandas as pd
import re
import os
import sys
import time
from openpyxl import load_workbook
from openpyxl import Workbook

def guardar_datos_en_excel(df_nuevo, archivo, hoja):
    """
    Guarda los datos de un DataFrame en una hoja específica de un archivo Excel.
    
    Si la hoja existe, agrega los datos sin sobrescribir. Si la hoja no existe, 
    la crea y guarda los datos. Si el archivo no existe, lo crea y guarda los datos.
    
    Args:
    - df_nuevo: DataFrame con los datos a agregar.
    - archivo: Ruta al archivo Excel donde se guardarán los datos.
    - hoja: Nombre de la hoja en el archivo Excel.
    """
    # Verificar si el archivo tiene la extensión .xlsx
    if not archivo.endswith('.xlsx'):
        raise ValueError(f"El archivo {archivo} no tiene la extensión .xlsx.")
    
    try:
        # Si el archivo no existe, crearlo
        if not os.path.exists(archivo):
            # Crear un nuevo archivo Excel si no existe
            libro = Workbook()
            libro.save(archivo)
            print(f"Archivo {archivo} creado exitosamente.")
        
        # Intentar cargar el archivo Excel
        libro = load_workbook(archivo)
        
        # Verificar si la hoja existe
        if hoja in libro.sheetnames:
            # Leer la hoja existente si ya existe
            df_existente = pd.read_excel(archivo, sheet_name=hoja)
        else:
            # Si la hoja no existe, crear una hoja vacía
            df_existente = pd.DataFrame()

    except Exception as e:
        raise RuntimeError(f"Error al abrir el archivo Excel: {e}")
    

    # Guardar el DataFrame actualizado en la misma hoja
    with pd.ExcelWriter(archivo, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        # Borrar la hoja antes de escribir los nuevos datos
        if hoja in libro.sheetnames:
            std = libro[hoja]
            libro.remove(std)
        
        # Escribir el DataFrame actualizado en la hoja específica
        df_nuevo.to_excel(writer, sheet_name=hoja, index=False)
    
    print(f"\nDatos guardados exitosamente en la hoja '{hoja}' de {archivo}.")
def imprimir_bienvenida():
    # Mensaje grande
    print("\n")
    print("***************************************")
    print("** Data Voley Information Processing **")
    print("**              System               **")
    print("***************************************")
    time.sleep(1)  # Pausa de 1 segundo para hacer la transición más agradable
    # Mensaje más pequeño
    print("\nBy C.Braun")
    print("\n")
    time.sleep(1)  # Pausa antes de continuar con el proceso
def extraer_numeros(texto):
    """
    Extrae todos los números de un string y los retorna como un string.

    Args:
    - texto: El texto de entrada.

    Returns:
    - Un string que contiene solo los números extraídos.
    """
    return ''.join(re.findall(r'\d+', texto))
def imprimir_barra_progreso(iteracion, total, longitud=40):
    progreso = iteracion / total
    barra = '█' * int(progreso * longitud)  # Genera una barra de progreso
    espacios = ' ' * (longitud - len(barra))  # Rellena el espacio restante
    porcentaje = progreso * 100  # Calcula el porcentaje
    # Muestra la barra y el porcentaje
    sys.stdout.write(f'\r[{barra}{espacios}] {porcentaje:.2f}%')
    sys.stdout.flush()
def agregar_columna_id(df, ID):
    # Agregamos la nueva columna con el valor único para todas las filas
    df.insert(0,"ID", ID)
    # Retornamos el DataFrame actualizado
    return df
def agregar_columna_equipo(df, team):
    # Agregamos la nueva columna con el valor único para todas las filas
    df.insert(1,"Equipo", team)
    # Retornamos el DataFrame actualizado
    return df
def leer_input(path):
    ruta_actual = os.path.dirname(__file__)
    nombre_archivo = path
    ruta_archivo = os.path.join(ruta_actual, nombre_archivo)

    try:
        with open(ruta_archivo, "r", encoding="utf-8") as archivo:
            contenido = archivo.read().strip()  # Remueve espacios en blanco extra
            return contenido
    except FileNotFoundError:
        print(f"El archivo '{nombre_archivo}' no fue encontrado en la carpeta del script.")
        exit()
    except Exception as e:
        print(f"Ocurrió un error al leer el archivo: {e}")
        exit()
def cargar_info(path):
    df = pd.read_csv(path, sep=';',encoding='latin1')
    file_paths = df["RUTA"].tolist()
    equipos1 = df["EQUIPO1"].tolist()
    equipos2 = df["EQUIPO2"].tolist()
    return file_paths,equipos1,equipos2
def info_test():
    file_paths = ["FECHA1/AA ESPINHO 1 X 3 SL BENFICA.pdf","FECHA1/VC VIANA 0 X 3 SPORTING CP .pdf","FECHA1/SD CALDAS 3 X 2 VITORIA SC.pdf","FECHA1/AJ FONTE BASTARDO 3 X 0 KAIROS.pdf","FECHA1/SC ESPINHO 3 X 2 CASTELO MAIA .pdf","FECHA1/G.C SANTO TIRSO 0 X 3 AAS MAMEDE.pdf","FECHA1/ESMORIZ GC 3 X 2 LEIXOES SC.pdf"]
    equipos1 = ["AA Espinho","VC Viana-Casa Peixoto","SC Caldas/Aki D'el Mar","AJ Fonte Bastardo","SC Espinho","G.C. Santo Tirso","Esmoriz GC"]
    equipos2 = ["SL Benfica","Sporting CP","Vitória SC","Clube Kairós","Castelo Maia GC","AAS Mamede","Leixões SC"]
    return file_paths,equipos1,equipos2
def eliminar_items_con_letras(lista):
    return [item for item in lista if not any(char.isalpha() for char in str(item))]
def extract_text_lines(file_path):
    lines = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                lines.extend(text.splitlines())
    return lines
def imprimir_datos(datos):
    for i in datos:
        print(i,len(i))
def preprocesar_datos_jugador(linea, headers):
    # Dividir línea en valores individuales
    datos = linea.replace("( ", "(").split()
    
    # Verificar y ajustar columna 'L' o 'C'
    if "L" in datos[0]:
        temp = datos[0].replace("L", "")
        datos[0] = temp
        datos.insert(1, "L")
    if "C" in datos[0]:
        temp = datos[0].replace("C", "")
        datos[0] = temp
        datos.insert(1, "C")

    if datos[1] != "L" and datos[1] != "C" and datos[1] != "totalizar" and datos[1] != "total":
        datos.insert(1, ".")

    # Contar nombres, excluyendo abreviaturas (como "A.")
    def es_nombre_valido(item):
        return item.isalpha() or (len(item) > 1 and item[-1] == '.' and item[:-1].isalpha())

    n_nombres = sum(1 for item in datos[2:] if es_nombre_valido(item))

    # Unir el tercer nombre si corresponde
    if n_nombres == 3:
        datos[3] = datos[3] + " " + datos[4]
        datos.pop(4)

    return datos
def rellenar_datos(datos, headers):
    for linea in datos:
        if len(linea) != len(headers):
            for i in range((- len(linea)+ len(headers))):
                if linea[1] == "totalizar" or linea[1]=="total":
                    linea.insert(2,".")
                else:
                    linea.insert(4,".")
    return datos
def cambiar_errOP(df1,df2):
    temp = df1["Error"]
    df1["Error"] = df2["Error"]
    df2["Error"] = temp
    return df1,df2
def procesar_linea(linea):
    match = re.search(r'\b[1-5]\b(?![\.,])', linea)
    if match:
        numero = match.group()
        indice = match.end()
        texto_restante = linea[indice:].strip()
        return numero + " " + texto_restante
    else:
        return "Not found"
def procesar_info_por_set(lineas):
    resultados = [["Set","Ser", "Ata", "BK", "Error", "Tot", "Err", "Pts", "Tot", "Err", "Pos%", "(Prf%)", "Tot", "Err", "Blo", "Pts", "Pts%", "Pts"]]
    for linea in lineas:
        match = re.search(r"Set.*", linea) 
        if match:
            temp = match.group().replace("( ","(").split(" ")
            temp[0] = temp[0]+" "+temp[1]
            temp.pop(1)
            resultados.append(temp) 
    return resultados
def calcular_score_total(df):
    score_e1 = 0
    score_e2 = 0
    for i in df:
        e1,e2 = i.split("-")
        score_e1 += int(e1)
        score_e2 += int(e2)
    return score_e1,score_e2
def extraer_recepcion_y_palabras(linea):
    palabras = linea.split()
    resultado = []
    
    for i in range(1, len(palabras)):
        if palabras[i].isdigit():  # Verifica si el elemento es un número
            resultado.append(f"{palabras[i-1]} {palabras[i]}")
    
    return resultado[0],resultado[1]
def extraer_puntos_y_palabras_previas(linea):
    palabras = linea.split()
    resultado = []
    
    for i in range(2, len(palabras)):  # Comienza desde el índice 2 para evitar errores en índices negativos
        if palabras[i].isdigit():  # Verifica si el elemento es un número
            resultado.append(f"{palabras[i-2]} {palabras[i-1]} {palabras[i]}")
    
    return resultado[0],resultado[1]
def clean_string2(temp):
    lista = temp.split(" ")
    lista.pop(3)
    lista.pop(3)
    lista.pop(3)
    lista.pop(3)
    lista.pop(3)
    lista.pop(3)
    lista.pop(3)
    lista.pop(3)
    temp = " ".join(lista)
    return temp
def clean_string(temp):
    lista = temp.split(" ")
    lista.pop(0)
    lista.pop(0)
    lista.pop(-1)
    lista.pop(-1)
    temp = " ".join(lista)
    return temp
def clean_string3(temp):
    lista = temp.split(" ")
    lista.pop(0)
    lista.pop(0)
    lista.pop(-1)
    lista.pop(0)
    lista.pop(-1)
    lista.pop(-1)
    lista.pop(-1)
    lista.pop(-1)
    lista.pop(-1)
    temp = " ".join(lista)
    return temp
def clean_string4(temp):
    lista = temp.split(" ")
    lista.pop(0)
    lista.pop(0)
    lista.pop(0)
    lista.pop(0)
    lista.pop(0)
    lista.pop(-1)
    lista.pop(-1)
    lista.pop(-1)
    lista.pop(-1)
    lista.pop(-1)
    temp = " ".join(lista)
    return temp 
def procesar_partidos_seguro(file_path, equipo1, equipo2, lista_final, final_headers):
    try:
        procesar_partidos(file_path, equipo1, equipo2, lista_final, final_headers)
    except Exception as e:
        print(f"Error procesando el archivo: {file_path}")
        print(f"Detalles del error: {e}")
def procesar_partidos(file_path,equipo1,equipo2,lista_final,final_headers):
    lineas_contenido = extract_text_lines(file_path)
    match_data = lineas_contenido[0:12]
    linea_inicio_data_general = next((i for i, linea in enumerate(match_data) if "Date" in linea or "Data" in linea), None)
    linea_equipo_local = next((i for i, linea in enumerate(match_data) if equipo1 in linea ), None)
    linea_equipo_visita = next((i for i, linea in enumerate(match_data) if equipo2 in linea ), None)
    sets_local = re.search(r'\d+',match_data[linea_equipo_local]).group()
    sets_visita = re.search(r'\d+',match_data[linea_equipo_visita]).group()
    if int(sets_local) == 3:
        ganador_local = 1
        ganador_visita = 0
    else:
        ganador_local = 0
        ganador_visita = 1
    total_sets = int(sets_local)+int(sets_visita)
    Match_id = re.search(r'\d+',match_data[linea_inicio_data_general-1]).group()
    Match_date = re.search(r'\b\d{2}/\d{2}/\d{4}\b', match_data[linea_inicio_data_general]).group()
    match_ID = Match_date+"_"+Match_id
    match_general_data = [["Set","Duración","Parcial 8","Parcial 16","Parcial 21","Score"]]
    for linea in match_data[linea_inicio_data_general:linea_inicio_data_general+5]:
        resultado = procesar_linea(linea)
        if resultado == "Not found":
            break
        match_general_data.append(resultado.split(" "))
    df_match = agregar_columna_id(pd.DataFrame(match_general_data[1:], columns=match_general_data[0]),match_ID)
    df_match["Duración"] = pd.to_numeric(df_match["Duración"], errors="coerce")  # Convierte valores no válidos a NaN
    df_match["Set"] = pd.to_numeric(df_match["Set"], errors="coerce")  # Convierte valores no válidos a NaN
    total_duracion_temp = df_match["Duración"].sum()
    total_duracion = str(total_duracion_temp//0.60).split(".")[0]+"."+str(total_duracion_temp%0.60).split(".")[1]
    score_e1,score_e2 = calcular_score_total(df_match["Score"])
    score_total = str(score_e1)+"-"+str(score_e2)
    break_line = equipo1+" "+equipo2
    equipo1_linea = next((i for i, linea in enumerate(lineas_contenido) if linea.replace(" Set Pontos Serviço Recepção Ataque BK","").strip() == equipo1), None)
    equipo2_linea = next((i for i, linea in enumerate(lineas_contenido) if linea.replace(" Set Pontos Serviço Recepção Ataque BK","").strip() == equipo2), None)
    equipo1_fin, equipo2_fin = [i for i, linea in enumerate(lineas_contenido) if ("Jogadores totalizar" in linea)or ("Players total" in linea)]
    break_linea = next((i for i, linea in enumerate(lineas_contenido) if break_line in linea), None)
    equipo1_data = lineas_contenido[equipo1_linea:equipo2_linea]
    equipo2_data = lineas_contenido[equipo2_linea:break_linea]
    predata1 = equipo1_data[1].split()
    equipo1_data[1] = ["Numero","Libero","Apellido","Nombre"] + predata1
    predata2 = equipo2_data[1].split()
    equipo2_data[1] = ["Numero","Libero","Apellido","Nombre"] + predata2
    datos_jugadores1 = [preprocesar_datos_jugador(linea,equipo1_data[1]) for linea in equipo1_data[2:equipo1_fin+1-equipo1_linea]]
    datos_por_set_e1 = procesar_info_por_set(equipo1_data[equipo1_fin+1-equipo1_linea:equipo1_fin-equipo1_linea+total_sets+2])
    datos_por_set_e2 = procesar_info_por_set(equipo2_data[equipo2_fin+1-equipo2_linea:equipo2_fin-equipo2_linea+total_sets+2])
    header_2= ['Set', 'Ser', 'Ata', 'BK', 'Error', 'Tot S', 'Err S', 'Pts S', 'Tot R', 'Err R', 'Pos% R', '(Prf%) R', 'Tot A', 'Err A', 'Blo A', 'Pts A', 'Pts% A', 'Pts BK']
    df_data_por_set_e1 = agregar_columna_equipo(agregar_columna_id(pd.DataFrame(datos_por_set_e1[1:],columns=header_2),match_ID),equipo1)
    df_data_por_set_e2 = agregar_columna_equipo(agregar_columna_id(pd.DataFrame(datos_por_set_e2[1:],columns=header_2),match_ID),equipo2)
    cambiar_errOP(df_data_por_set_e1,df_data_por_set_e2)
    headers_1 = ['Numero', 'Libero', 'Apellido', 'Nombre', '1', '2', '3', '4',"5", 'Voto', 'Tot', 'BP', 'V-P', 'Tot S', 'Err S', 'Pts S', 'Tot R', 'Err R', 'Pos% R', '(Prf%) R', 'Tot A', 'Err A', 'Blo A', 'Pts A', 'Pts% A', 'Pts BK']
    datos1 = rellenar_datos(datos_jugadores1,headers_1)
    datos_jugadores2 = [preprocesar_datos_jugador(linea,equipo2_data[1]) for linea in equipo2_data[2:equipo2_fin+1-equipo2_linea]]
    datos2 = rellenar_datos(datos_jugadores2,headers_1)
    df1 = agregar_columna_equipo(agregar_columna_id(pd.DataFrame(datos1, columns=headers_1),match_ID),equipo1)
    sub_df1 = df1["Voto"].copy()
    sub_df1 = pd.to_numeric(sub_df1, errors='coerce')
    promedio1 = sub_df1.mean()
    df2 = agregar_columna_equipo(agregar_columna_id(pd.DataFrame(datos2, columns=headers_1),match_ID),equipo2)
    sub_df2 = df2["Voto"].copy()
    sub_df2 = pd.to_numeric(sub_df2, errors='coerce')
    promedio2 = sub_df2.mean()
    # print("-----------------")
    # for i in lineas_contenido[break_linea:]:
    #     print(i)
    # print("-----------------")
    data = lineas_contenido[break_linea:]
    rece1,rece2 = extraer_recepcion_y_palabras(data[2])
    puntos1,puntos2 = extraer_puntos_y_palabras_previas(data[3])
    if any(char.isdigit() for char in data[10]):
        temp1 = data[10] 
        temp2 = data[11]
    else:
        temp1 = data[11]
        temp2 = data[12]
    # if len(data[6].split(" "))<=8:
    #     temp3 = clean_string(data[7])
    # else:
    #     temp3 = clean_string3(data[6])
    serve1,serve2 = extraer_recepcion_y_palabras(clean_string(temp1))
    ace1 , ace2 = extraer_puntos_y_palabras_previas(clean_string2(clean_string(temp2)))
    # attack_positive = eliminar_items_con_letras([temp3.split(" ")])
    # attack_exclamative = eliminar_items_con_letras([clean_string4(temp2).split(" ")])
    # contra = [data[15].split(" ")]
    # headers = ["Err1", "Blo1", "Pts%s1","Tot1", "Tot2", "Pts%s2", "Blo2", "Err2"]
    headers2 = ["Equipo","Recepcion","Puntos SO","Servicio","Puntos BP"]
    temp = [[equipo1,extraer_numeros(rece1),extraer_numeros(puntos1),extraer_numeros(serve1),extraer_numeros(ace1)],[equipo2,extraer_numeros(rece2),extraer_numeros(puntos2),extraer_numeros(serve2),extraer_numeros(ace2)]]
    # df_positive = pd.DataFrame(attack_positive,columns=headers)
    # df_exclamative = pd.DataFrame(attack_exclamative,columns=headers)
    # df_contra = pd.DataFrame(contra,columns=headers)
    df_metricas = agregar_columna_id(pd.DataFrame(temp,columns=headers2),match_ID)
    # headers_3 = ["ID","Equipo1","Equipo2","Numero de Partido","Fecha","Sets Local","Sets Visita","Set Total","Duración","Score Acumulado"]
    match_info = [[match_ID,equipo1,equipo2,Match_id,Match_date,sets_local,sets_visita,str(int(sets_local)+int(sets_visita)),total_duracion,str(score_e1)+"-"+str(score_e2)]]
    # df_general = pd.DataFrame(match_info,columns=headers_3)
    final_headers = ["Partido","Fecha","Equipo","Ganador","Local","Sets","Duracion","Resultado","Eva","Tot Puntos","BP Puntos","G-P Puntos","Tot Saque","Err Saque","Pts Saque","Tot Recepcion","Err Recepcion","Pos% Recepcion","(Exc%) Recepcion","Tot Ataque","Err Ataque","BL Ataque","Pts Ataque","Pts% Ataque", "Pts BL","Recepciones por Puntos","Saques por Break Point" ]
    equipo1_datos = [Match_id,Match_date,equipo1,ganador_local,1,sets_local,total_duracion,score_total,float(promedio1)]
    equipo1_datos.extend(df1.iloc[-1].to_list()[-16:])
    equipo1_datos.extend([int(df_metricas["Recepcion"][0])/int(df_metricas["Puntos SO"][0]),int(df_metricas["Servicio"][0])/int(df_metricas["Puntos BP"][0])])
    equipo2_datos = [Match_id,Match_date,equipo2,ganador_visita,0,sets_visita,total_duracion,score_total,float(promedio2)]
    equipo2_datos.extend(df2.iloc[-1].to_list()[-16:])
    equipo2_datos.extend([int(df_metricas["Recepcion"][1])/int(df_metricas["Puntos SO"][1]),int(df_metricas["Servicio"][1])/int(df_metricas["Puntos BP"][1])])
    lista_final.append(equipo1_datos)
    lista_final.append(equipo2_datos)

    # print("---------------------")
    # print(df_final)
    # guardar_datos_en_excel(df_final,"DATOS.xlsx","Datos")
    # print("---------------------")

    # print("PARTIDO "+equipo1.upper()+" v/s "+ equipo2.upper())
    # print(df_general)

    # print(df_match)
    # guardar_datos_en_excel(df_match,"DATOS.xlsx","Sets_General")
    # print("Duración total: "+ str(total_duracion))
    # print("Score Total: "+str(score_e1)+"-"+str(score_e2))
    # print("______________________________________________")
    # print("EQUIPO "+equipo1.upper())


    # guardar_datos_en_excel(df1,"DATOS.xlsx","Jugadores")
    # print("______________________________________________")
    # print(df_data_por_set_e1)
    # guardar_datos_en_excel(df_data_por_set_e1,"DATOS.xlsx","Por_Set")
    # print("______________________________________________")
    # print("EQUIPO "+equipo2.upper())
    # print(df2)
    # guardar_datos_en_excel(df2,"DATOS.xlsx","Jugadores")
    # print("______________________________________________")
    # print(df_data_por_set_e2)
    # guardar_datos_en_excel(df_data_por_set_e2,"DATOS.xlsx","Por_Set")
    # print("______________________________________________")
    # print("______________________________________________")

    # print(df_metricas)
    # guardar_datos_en_excel(df_metricas,"DATOS.xlsx","Recepción_Servicio")
    # print("-----------------")
    # print("1° ATAQUE RECE(+#)")
    # print(df_positive)
    # print("1° ATAQUE RECE(-!)")
    # print(df_exclamative)
    # print("CONTRA ATAQUE")
    # print(df_contra)
    # print("-----------------")

#file_paths,equipos1,equipos2 = info_test()
leer_input("ruta_csv.txt")
file_paths,equipos1,equipos2 = cargar_info(leer_input("ruta_csv.txt"))
lista_final, final_headers = [],["Partido","Fecha","Equipo","Ganador","Local","Sets","Duracion","Resultado","Eva","Tot Puntos","BP Puntos","G-P Puntos","Tot Saque","Err Saque","Pts Saque","Tot Recepcion","Err Recepcion","Pos% Recepcion","(Exc%) Recepcion","Tot Ataque","Err Ataque","BL Ataque","Pts Ataque","Pts% Ataque", "Pts BL","Recepciones por Puntos","Saques por Break Point" ]
total = len(file_paths)
imprimir_bienvenida()
for i in range(total):
    procesar_partidos_seguro(file_paths[i],equipos1[i],equipos2[i], lista_final,final_headers)
    imprimir_barra_progreso(i + 1, total)
guardar_datos_en_excel(pd.DataFrame(lista_final,columns=final_headers),"DATOS.xlsx","Datos")